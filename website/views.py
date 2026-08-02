# website/views.py
from django.views.generic import TemplateView, ListView
from django.views import View
from .models import MotorComment, Title, RaceDay, Event
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseNotAllowed, HttpResponse, HttpResponseNotFound
import io
import json
from urllib.parse import quote
from datetime import date, datetime, time, timedelta
from django.shortcuts import get_object_or_404
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.db import transaction
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework import viewsets
from .serializers import RaceDaySerializer, EventSerializer
import openpyxl


@method_decorator(ensure_csrf_cookie, name="dispatch")
class Motor_Comments(TemplateView):
    template_name = "website/motor_comments.html"

    # 常に選択肢の先頭に出す固定タイトル
    FIRST_TITLE = "ボートの時間ご視聴ありがとう競走"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        titles = list(Title.objects.values_list("title", flat=True))
        # 固定タイトルを先頭に。DB側に同名があっても重複させない
        context["titles"] = [self.FIRST_TITLE] + [t for t in titles if t != self.FIRST_TITLE]
        return context


@method_decorator(ensure_csrf_cookie, name="dispatch")
class Calendar(TemplateView):
    template_name = "website/calendar.html"


class MotorCommentListCreateAPI(View):
    """
    GET  /api/machines/<machine_no>/posts  → その号機のコメント一覧(JSON)
    POST /api/machines/<machine_no>/posts  → 1件作成(JSON; 201)
    """
    def get(self, request, machine_no):
        qs = MotorComment.objects.filter(machine_no=machine_no)
        data = [
            {
                "id": c.id,
                "author": c.author or "匿名",
                "racer": c.racer or "",
                "content": c.content,
                "scheduled_at": c.scheduled_at.isoformat() if c.scheduled_at else None,
                "created_at": c.created_at.isoformat(),
                "title": c.title or "",
                "boat_no": c.boat_no or "",
                "parts_exchange": c.parts_exchange or "",
            }
            for c in qs
        ]
        return JsonResponse(data, safe=False)

    def post(self, request, machine_no):
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except json.JSONDecodeError:
            return HttpResponseBadRequest("invalid json")

        content = (payload.get("content") or "").strip()
        parts_exchange = (payload.get("parts_exchange") or "").strip()
        # 部品交換の入力があれば本文は空欄でもよい
        if not content and not parts_exchange:
            return HttpResponseBadRequest("content is required")

        boat_no = (str(payload.get("boat_no") or "")).strip()
        if boat_no and not (boat_no.isdigit() and len(boat_no) <= 3):
            return HttpResponseBadRequest("boat_no must be a number of up to 3 digits")

        author = (payload.get("author") or "匿名").strip() or "匿名"
        racer = (payload.get("racer") or "").strip()
        scheduled = payload.get("scheduled_at")
        title = (payload.get("title") or "").strip()

        obj = MotorComment(machine_no=machine_no, author=author, content=content, racer=racer, title=title,
                           boat_no=boat_no, parts_exchange=parts_exchange)
        if scheduled:
            try:
                obj.scheduled_at = date.fromisoformat(scheduled)
            except ValueError:
                pass
        obj.save()

        return JsonResponse({
            "id": obj.id,
            "author": obj.author,
            "racer": obj.racer,
            "content": obj.content,
            "scheduled_at": obj.scheduled_at.isoformat() if obj.scheduled_at else None,
            "created_at": obj.created_at.isoformat(),
            "title": obj.title,
            "boat_no": obj.boat_no or "",
            "parts_exchange": obj.parts_exchange or "",
        }, status=201)


class MotorCommentDetailAPI(View):
    """
    DELETE /api/machines/<machine_no>/posts/<pk>         → 削除
    PUT    /api/machines/<machine_no>/posts/<pk>         → 更新
    POST   /api/machines/<machine_no>/posts/<pk>/delete  → 削除（互換）
    POST   /api/machines/<machine_no>/posts/<pk>/update  → 更新（互換）
    """
    def delete(self, request, machine_no, pk):
        obj = get_object_or_404(MotorComment, pk=pk, machine_no=machine_no)
        obj.delete()
        return HttpResponse(status=204)

    def get(self, *args, **kwargs):
        return HttpResponseNotAllowed(["DELETE", "PUT"])

    def put(self, request, machine_no, pk):
        obj = get_object_or_404(MotorComment, pk=pk, machine_no=machine_no)
        return self._apply_update(obj, request)

    def post(self, request, machine_no, pk):
        path = str(request.path)
        # /.../delete と /.../update だけ許可（メソッドを弾く環境向けの互換ルート）
        if path.endswith("/delete"):
            obj = get_object_or_404(MotorComment, pk=pk, machine_no=machine_no)
            obj.delete()
            return HttpResponse(status=204)
        if path.endswith("/update"):
            obj = get_object_or_404(MotorComment, pk=pk, machine_no=machine_no)
            return self._apply_update(obj, request)
        return HttpResponseNotAllowed(["DELETE", "PUT"])

    def _apply_update(self, obj, request):
        """既存の MotorComment を payload で上書き保存する（POST 作成と同じ検証）"""
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except json.JSONDecodeError:
            return HttpResponseBadRequest("invalid json")

        content = (payload.get("content") or "").strip()
        parts_exchange = (payload.get("parts_exchange") or "").strip()
        if not content and not parts_exchange:
            return HttpResponseBadRequest("content is required")

        boat_no = (str(payload.get("boat_no") or "")).strip()
        if boat_no and not (boat_no.isdigit() and len(boat_no) <= 3):
            return HttpResponseBadRequest("boat_no must be a number of up to 3 digits")

        obj.author = (payload.get("author") or "匿名").strip() or "匿名"
        obj.racer = (payload.get("racer") or "").strip()
        obj.content = content
        obj.title = (payload.get("title") or "").strip()
        obj.boat_no = boat_no
        obj.parts_exchange = parts_exchange
        scheduled = payload.get("scheduled_at")
        if scheduled:
            try:
                obj.scheduled_at = date.fromisoformat(scheduled)
            except ValueError:
                pass
        else:
            obj.scheduled_at = None
        obj.save()

        return JsonResponse({
            "id": obj.id,
            "author": obj.author,
            "racer": obj.racer,
            "content": obj.content,
            "scheduled_at": obj.scheduled_at.isoformat() if obj.scheduled_at else None,
            "created_at": obj.created_at.isoformat(),
            "title": obj.title,
            "boat_no": obj.boat_no or "",
            "parts_exchange": obj.parts_exchange or "",
        }, status=200)


@method_decorator(ensure_csrf_cookie, name="dispatch")
class MotorCommentDetailView(TemplateView):
    template_name = "website/motor_comments_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        machine_no = self.kwargs.get("machine_no")

        if not (1 <= machine_no <= 100):
            raise HttpResponseNotFound("指定された号機は存在しません")

        context["machine_no"] = machine_no
        context["titles"] = list(Title.objects.values_list("title", flat=True))
        return context


class RaceDayViewSet(viewsets.ModelViewSet):
    queryset = RaceDay.objects.all()
    serializer_class = RaceDaySerializer


class EventViewSet(viewsets.ModelViewSet):
    queryset = Event.objects.all()
    serializer_class = EventSerializer

    def create(self, request, *args, **kwargs):
        print("受信データ:", request.data)
        return super().create(request, *args, **kwargs)


class Motor_Comments_Index(TemplateView):
    template_name = 'website/motor_comments_index.html'


class Motor_Comments_Total(ListView):
    model = MotorComment
    template_name = 'website/motor_comments_total.html'
    context_object_name = 'liston'


from .scrape1 import motor_scrape
URL = 'https://www.boatrace-suminoe.jp/asp/suminoe/contents/01history/ranking_motor.php'


def grid_data_api(request):
    import pandas as pd
    df = motor_scrape(URL)
    df['ratio'] = pd.to_numeric(df['ratio'], errors='coerce')
    top6_list = df.nlargest(6, 'ratio')['number'].tolist()
    data = {
        "machine_numbers": [i for i in df['number']],
        "display_values": [i for i in df['ratio']],
        "top6": top6_list,
    }
    return JsonResponse(data)


# ============================================================
# Excelからコメントを一括取込（コメントの一部をまとめて追加する補助機能）
# ------------------------------------------------------------
# ・追加（新規作成）専用。既存データの削除・上書きは一切しない。
# ・部品交換(parts_exchange)は取込対象外（手入力を継続）。
# ・「発言日」列 → created_at に反映（auto_now_add を保存後 update で上書き）。
#   コメントの流れは created_at の時系列で追跡するため。scheduled_at(入力日)は使わない。
# ・厳格な all-or-nothing 検証: 1行でも不正なら何も保存しない。
# ============================================================
def _xl_str(v):
    """openpyxl のセル値を文字列へ。数値の 24.0 は "24"、日付は ISO 文字列にする。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _xl_int(v):
    if v is None or v == "":
        return None
    try:
        return int(str(v).strip()) if isinstance(v, str) else int(v)
    except (ValueError, TypeError):
        return None


def _xl_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().replace("/", "-")
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


# Excelヘッダー（列順）と各列の説明。取込・テンプレDLの両方で共有する。
MOTOR_EXCEL_HEADERS = ["号機", "コメント", "発言日", "開催", "使用選手", "使用ボート", "投稿者"]
MOTOR_EXCEL_NOTES = {
    "号機": "必須：号機番号（正の整数）",
    "コメント": "必須：コメント本文",
    "発言日": "必須：発言された日。時系列の並びに使用（例 2026-07-10）",
    "開催": "任意：セルのドロップダウンから選択（完全一致必須）",
    "使用選手": "任意：10文字まで",
    "使用ボート": "任意：数字3桁まで",
    "投稿者": "任意：空欄なら「匿名」",
}


def motor_titles_ordered():
    """開催ドロップダウン用の順序付きリスト（先頭に固定タイトル、前後空白除去）。"""
    titles = list(Title.objects.values_list("title", flat=True))
    ordered = [Motor_Comments.FIRST_TITLE] + [t for t in titles if t != Motor_Comments.FIRST_TITLE]
    return [t.strip() for t in ordered]


def build_motor_template_workbook():
    """現在の開催一覧を反映したコメント取込テンプレート(openpyxl Workbook)を生成する。"""
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    valid_titles = motor_titles_ordered()
    max_rows = 500
    widths = {"号機": 8, "コメント": 46, "発言日": 14, "開催": 40,
              "使用選手": 12, "使用ボート": 12, "投稿者": 12}

    wb = Workbook()
    ws = wb.active
    ws.title = "入力"
    header_fill = PatternFill("solid", fgColor="2F80ED")
    header_font = Font(color="FFFFFF", bold=True)
    for c, name in enumerate(MOTOR_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(MOTOR_EXCEL_NOTES[name], "template")
        ws.column_dimensions[get_column_letter(c)].width = widths[name]
    ws.freeze_panes = "A2"

    # 開催リストは別シート（非表示）に置き、ドロップダウンで参照（255字制限回避）
    ws_list = wb.create_sheet("リスト")
    ws_list.sheet_state = "hidden"
    for i, t in enumerate(valid_titles, start=1):
        ws_list.cell(row=i, column=1, value=t)
    list_ref = f"リスト!$A$1:$A${len(valid_titles)}"

    kaisai_col = get_column_letter(MOTOR_EXCEL_HEADERS.index("開催") + 1)
    dv_title = DataValidation(type="list", formula1=f"={list_ref}", allow_blank=True,
                              showErrorMessage=True)
    dv_title.error = "リストから選択してください（開催名は完全一致が必要です）"
    dv_title.errorTitle = "開催名が不正"
    dv_title.prompt = "ドロップダウンから開催を選択"
    dv_title.promptTitle = "開催"
    ws.add_data_validation(dv_title)
    dv_title.add(f"{kaisai_col}2:{kaisai_col}{max_rows + 1}")

    boat_col = get_column_letter(MOTOR_EXCEL_HEADERS.index("使用ボート") + 1)
    dv_boat = DataValidation(type="whole", operator="between", formula1="0", formula2="999",
                             allow_blank=True, showErrorMessage=True)
    dv_boat.error = "0〜999の数字で入力してください"
    dv_boat.errorTitle = "使用ボートが不正"
    ws.add_data_validation(dv_boat)
    dv_boat.add(f"{boat_col}2:{boat_col}{max_rows + 1}")

    mno_col = get_column_letter(MOTOR_EXCEL_HEADERS.index("号機") + 1)
    dv_mno = DataValidation(type="whole", operator="greaterThan", formula1="0",
                            allow_blank=True, showErrorMessage=True)
    dv_mno.error = "1以上の整数で入力してください"
    dv_mno.errorTitle = "号機が不正"
    ws.add_data_validation(dv_mno)
    dv_mno.add(f"{mno_col}2:{mno_col}{max_rows + 1}")

    date_col = get_column_letter(MOTOR_EXCEL_HEADERS.index("発言日") + 1)
    for r in range(2, max_rows + 2):
        ws[f"{date_col}{r}"].number_format = "yyyy-mm-dd"

    return wb


@method_decorator(ensure_csrf_cookie, name="dispatch")
class MotorCommentExcelUpload(TemplateView):
    template_name = "website/motor_comments_upload.html"

    # Excelヘッダー → 用途
    HEADER_MAP = {
        "号機": "machine_no",
        "コメント": "content",
        "発言日": "created_at",
        "使用選手": "racer",
        "開催": "title",
        "投稿者": "author",
        "使用ボート": "boat_no",
    }
    REQUIRED_HEADERS = ["号機", "コメント", "発言日"]

    def _valid_titles(self):
        # 入力フォームの開催ドロップダウン／テンプレのドロップダウンと同一の一覧
        # （完全一致の突合先）。取込値は _xl_str で前後空白を除去して比較するため、
        # motor_titles_ordered() 側も strip 済み（内部の改行は保持）。
        return set(motor_titles_ordered())

    def _error(self, details, **kwargs):
        context = self.get_context_data(**kwargs)
        context["error"] = "エラーが発生しました。ファイルを見直してください。"
        context["error_details"] = details
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        f = request.FILES.get("excel")
        if not f:
            return self._error(["ファイルが選択されていません。"], **kwargs)
        if not f.name.lower().endswith((".xlsx", ".xlsm")):
            return self._error(["拡張子が .xlsx / .xlsm のExcelを選んでください。"], **kwargs)

        wb = None
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            ws = wb.active
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        except Exception:
            return self._error(["Excelファイルとして読み込めませんでした。"], **kwargs)
        finally:
            if wb is not None:
                wb.close()

        if not rows:
            return self._error(["データがありません。"], **kwargs)

        # ヘッダー行から列位置を特定
        header = [_xl_str(c) for c in rows[0]]
        col_index = {h: i for i, h in enumerate(header) if h in self.HEADER_MAP}
        missing = [h for h in self.REQUIRED_HEADERS if h not in col_index]
        if missing:
            return self._error([f"必須列が見つかりません: {', '.join(missing)}"], **kwargs)

        valid_titles = self._valid_titles()
        errors = []
        parsed = []
        for r_idx, raw in enumerate(rows[1:], start=2):  # Excelの行番号（ヘッダー=1行目）
            def cell(h):
                i = col_index.get(h)
                return raw[i] if (i is not None and i < len(raw)) else None

            # 完全な空行はスキップ
            if all(_xl_str(c) == "" for c in raw):
                continue

            row_err = []
            machine_no = _xl_int(cell("号機"))
            if machine_no is None or machine_no <= 0:
                row_err.append("号機が不正（正の整数が必要）")

            content = _xl_str(cell("コメント"))
            if not content:
                row_err.append("コメントが空")

            said = _xl_date(cell("発言日"))
            if said is None:
                row_err.append("発言日が不正（日付が必要）")

            title = _xl_str(cell("開催"))
            if title and title not in valid_titles:
                row_err.append(f"開催『{title}』が開催リストと一致しない")

            boat_no = _xl_str(cell("使用ボート"))
            if boat_no and not (boat_no.isdigit() and len(boat_no) <= 3):
                row_err.append("使用ボートは数字3桁まで")

            racer = _xl_str(cell("使用選手"))
            if len(racer) > 10:
                row_err.append("使用選手は10文字まで")

            author = _xl_str(cell("投稿者")) or "匿名"
            if len(author) > 50:
                row_err.append("投稿者は50文字まで")

            if row_err:
                errors.append(f"{r_idx}行目: " + " / ".join(row_err))
            else:
                parsed.append({
                    "machine_no": machine_no, "content": content, "racer": racer,
                    "title": title, "author": author, "boat_no": boat_no, "said": said,
                })

        if errors:
            return self._error(errors, **kwargs)
        if not parsed:
            return self._error(["登録対象の行がありません。"], **kwargs)

        # all-or-nothing 保存（新規追加のみ・既存データは触らない）
        try:
            with transaction.atomic():
                for i, row in enumerate(parsed):
                    obj = MotorComment(
                        machine_no=row["machine_no"], author=row["author"],
                        racer=row["racer"], content=row["content"], title=row["title"],
                        boat_no=row["boat_no"], parts_exchange="", scheduled_at=None,
                    )
                    obj.save()
                    # created_at を発言日で上書き（auto_now_add を回避）。
                    # 時刻は正午(JST)にする: フロント(motor_comments_detail.js)は
                    # created_at.isoformat() を slice(0,10) で日付として使い、APIはUTC表現の
                    # ISOを返すため。真夜中だとUTCで前日になり1日ずれる。正午なら
                    # UTC(03:00)でも同じ暦日を保てる。
                    # 同日内はExcelの行順を保つため秒をずらして順序を安定させる。
                    dt = timezone.make_aware(datetime.combine(row["said"], time(12, 0))) + timedelta(seconds=i)
                    MotorComment.objects.filter(pk=obj.pk).update(created_at=dt)
        except Exception as e:
            return self._error([f"保存中にエラーが発生しました: {e}"], **kwargs)

        context = self.get_context_data(**kwargs)
        context["result"] = f"{len(parsed)}件のコメントを登録しました。"
        return self.render_to_response(context)


class MotorCommentTemplateDownload(View):
    """GET /website/motor_comments_template/ → 現在の開催一覧入りの取込テンプレxlsxを配布。

    開催(Title)は随時差し替わるため、毎回その時点の一覧でドロップダウンを生成する
    （静的ファイルとして持たず陳腐化を防ぐ）。
    """
    def get(self, request):
        wb = build_motor_template_workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        fname = "モーターコメント取込テンプレート.xlsx"
        # 日本語ファイル名は RFC5987 の filename* で渡す（ASCIIフォールバックも併記）
        resp["Content-Disposition"] = (
            "attachment; filename=motor_comment_template.xlsx; "
            f"filename*=UTF-8''{quote(fname)}"
        )
        return resp
